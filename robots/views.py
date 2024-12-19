import json
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.utils.dateparse import parse_datetime
from .models import Robot
import datetime
from django.db.models import Count
from openpyxl import Workbook

class RobotCreateView(View):
    def post(self, request, *args, **kwargs):
        try:
            body = json.loads(request.body)
            model = body.get("model")
            version = body.get("version")
            created = body.get("created")

            if not model or not version or not created:
                return JsonResponse({"error": "Missing required fields"}, status=400)

            if not model.isalnum() or len(model) != 2:
                return JsonResponse({"error": "Invalid model format"}, status=400)

            if not version.isalnum() or len(version) != 2:
                return JsonResponse({"error": "Invalid version format"}, status=400)

            if not Robot.objects.filter(model=model).exists():
                return JsonResponse({"error": "Model does not exist in the system"}, status=400)

            try:
                created_date = parse_datetime(created)
                if not created_date:
                    raise ValueError
            except ValueError:
                return JsonResponse({"error": "Invalid created format"}, status=400)

            serial = f"{model}-{version}"

            robot = Robot.objects.create(
                serial=serial,
                model=model,
                version=version,
                created=created_date,
            )
            return JsonResponse(
                {
                    "serial": robot.serial,
                    "model": robot.model,
                    "version": robot.version,
                    "created": robot.created.isoformat(),
                },
                status=201,
            )
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


class WeeklyProductionSummaryView(View):
    def get(self, request, *args, **kwargs):
        end_date = datetime.datetime.now()
        start_date = end_date - datetime.timedelta(days=7)

        data = (
            Robot.objects.filter(created__range=[start_date, end_date])
            .values("model", "version")
            .annotate(count=Count("id"))
            .order_by("model", "version")
        )

        workbook = Workbook()
        current_model = None
        worksheet = None

        for row in data:
            model = row["model"]
            version = row["version"]
            count = row["count"]

            if model != current_model:
                worksheet = workbook.create_sheet(title=model)
                worksheet.append(["Модель", "Версия", "Количество за неделю"])
                current_model = model

            worksheet.append([model, version, count])

        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="weekly_production_summary.xlsx"'
        workbook.save(response)
        return response
